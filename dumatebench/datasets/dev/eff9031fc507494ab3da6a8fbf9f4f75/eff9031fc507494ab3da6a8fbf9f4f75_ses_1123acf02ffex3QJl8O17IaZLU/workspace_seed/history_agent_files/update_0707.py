import openpyxl, re

src = '/home/work/dumate/eff9031fc507494ab3da6a8fbf9f4f75/workspace/ses_1123acf02ffex3QJl8O17IaZLU/塑壳领料单_20260704.xlsx'
wb = openpyxl.load_workbook(src)
ws = wb['塑壳领料单']

def nn(v):
    if v is None: return None
    s = str(int(v)) if isinstance(v, float) and v == int(v) else str(v)
    return re.sub(r'^[Mm][Oo]', '', s)

def nq(v):
    if v is None: return None
    return int(v) if isinstance(v, float) and v == int(v) else v

existing = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    rd = list(row[:7])
    if rd[1] is not None:
        rd[2] = nn(rd[2]); rd[3] = nq(rd[3])
        existing.append(rd)

enos = set()
for r in existing:
    if r[2]: enos.add(str(r[2]).strip())

print(f"Existing: {len(existing)}")

new_entries = [
    # 张培 7/7 07:54 - EV3 P02373D04500550JF
    ['EV3', 'P02373D04500550JF', '840104', 200, '13点用', '7.7', ''],
    ['EV3', 'P02373D04500550JF', '840105', 200, '16点用', '7.7', ''],
    ['EV3', 'P02373D04500550JF', '840106', 200, '晚上加班用', '7.7', ''],
    
    # 秦玉英 7/7 07:56 - 比亚迪八合一B P05335D05500450JB, 840424=1 (多退领出)
    ['比亚迪八合一B', 'P05335D05500450JB', '840424', 1, '20点送', '7.7', '多退领出'],
    
    # 姚献蓉 7/7 08:01 - 智新PA2468D05500350JC, 840294=160, 840295=160, 明天上午用
    ['智新', 'PA2468D05500350JC', '840294', 160, '明天上午用', '7.7', ''],
    ['智新', 'PA2468D05500350JC', '840295', 160, '明天上午用', '7.7', ''],
    
    # 周素芝 7/7 08:11 - 比亚迪P05335D08000900JB
    ['比亚迪', 'P05335D08000900JB', '838016', 30, '白班下午4点左右用', '7.7', '补领'],
    ['比亚迪', 'P05335D08000900JB', '844900', 200, '白班下午4点左右用', '7.7', ''],
    ['比亚迪', 'P05335D08000900JB', '844901', 200, '夜班用', '7.7', ''],
    ['比亚迪', 'P05335D08000900JB', '844902', 200, '夜班用', '7.7', ''],
    ['比亚迪', 'P05335D08000900JB', '844903', 200, '夜班用', '7.7', ''],
    ['比亚迪', 'P05335D08000900JB', '844904', 200, '明天早上用', '7.7', ''],
    
    # 秦玉英 7/7 08:29 - 伊控(GN3) PB0033D05000800JA, MO839599→839599=200
    ['伊控(GN3)', 'PB0033D05000800JA', '839599', 200, '17点后先送100其余开班要', '7.7', ''],
    
    # 李鸿 7/7 08:30 - 广汽多合一500/500JF
    ['广汽多合一', '500/500JF', '843129', 200, '夜班', '7.7', ''],
    ['广汽多合一', '500/500JF', '844980', 200, '夜班', '7.7', ''],
    ['广汽多合一', '500/500JF', '844981', 200, '夜班', '7.7', ''],
    ['广汽多合一', '500/500JF', '844983', 200, '夜班', '7.7', ''],
    
    # 李鸿 7/7 08:41 - 依控PB0033D09200500JA, 845078=120, 明天白班
    ['依控', 'PB0033D09200500JA', '845078', 120, '明天白班', '7.7', ''],
    
    # 李鸿 7/7 13:42 - 法雷奥PA2447D05000810JA, MO842371→842371=100, 明天早上8:00
    ['法雷奥', 'PA2447D05000810JA', '842371', 100, '明天早上8:00', '7.7', ''],
    
    # 熊冬 7/7 14:35 - P09561D05000450JA
    [None, 'P09561D05000450JA', '843004', 200, '18点用', '7.7', ''],
    [None, 'P09561D05000450JA', '843005', 200, '夜班用', '7.7', ''],
    [None, 'P09561D05000450JA', '845156', 200, '夜班用', '7.7', ''],
    
    # 宗唤唤 7/7 14:54 - EV5线t2g降容P09561D05000730Ja
    ['EV5线/t2g降容', 'P09561D05000730Ja', '845928', 200, '明天开班用', '7.7', ''],
    ['EV5线/t2g降容', 'P09561D05000730Ja', '845929', 200, '明天上午十一点左右用', '7.7', ''],
]

truly_new = []
for e in new_entries:
    no = str(e[2]).strip()
    if no not in enos:
        truly_new.append(e)
        print(f"  + {e}")
    else:
        print(f"  dup: {no}")

print(f"\nNew: {len(truly_new)}")
all_data = existing + truly_new

for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for c in row: c.value = None

h = ['产品系列', '物料编码/规格', '单号', '数量', '需要时间', '发单时间', '备注']
for i, x in enumerate(h, 1): ws.cell(row=1, column=i, value=x)
for idx, rd in enumerate(all_data, 2):
    for ci, v in enumerate(rd, 1): ws.cell(row=idx, column=ci, value=v)

wb.save(src)

tc = {}
for rd in all_data:
    t = str(rd[5]).strip()
    if t: tc[t] = tc.get(t, 0) + 1
print(f"\nSaved: {src}")
for k in sorted(tc.keys()): print(f'  {k}: {tc[k]}条')
print(f'Total: {len(all_data)}')
mo = sum(1 for r in all_data if r[2] and 'MO' in str(r[2]).upper()[:2])
print(f'Mo前缀: {mo}')
