"""Summarise Excel artifacts for reference-free LLM judging.

The summary is intentionally evidence-oriented rather than exhaustive. Spreadsheet
benchmarks repeatedly show that value-only checks miss maintainability,
formatting, layout, formula, and robustness failures; this module extracts the
observable workbook signals a judge can cite without needing a gold answer.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}
OPENPYXL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def summarize_artifacts(
    artifact_dir: str | Path,
    *,
    max_files: int = 80,
    max_sheets: int = 12,
    max_cells_per_sheet: int = 80,
    max_formulas_per_sheet: int = 60,
    max_scan_rows: int = 200,
    max_scan_cols: int = 60,
) -> dict[str, Any]:
    """Return a JSON-serialisable evidence summary for an artifact directory."""
    root = Path(artifact_dir).expanduser().resolve()
    if not root.exists():
        raise FileNotFoundError(f"artifact directory does not exist: {root}")
    if not root.is_dir():
        raise NotADirectoryError(f"artifact path is not a directory: {root}")

    files = _iter_files(root, max_files=max_files)
    excel_files = [f for f in files if f["suffix"].lower() in EXCEL_EXTENSIONS]
    workbooks = [
        _summarize_workbook(
            root / f["relative_path"],
            root,
            max_sheets=max_sheets,
            max_cells_per_sheet=max_cells_per_sheet,
            max_formulas_per_sheet=max_formulas_per_sheet,
            max_scan_rows=max_scan_rows,
            max_scan_cols=max_scan_cols,
        )
        for f in excel_files
    ]

    return {
        "artifact_dir": str(root),
        "file_inventory": files,
        "inventory_truncated": len(files) >= max_files,
        "excel_file_count": len(excel_files),
        "non_excel_file_count": len(files) - len(excel_files),
        "workbooks": workbooks,
        "aggregate_signals": _aggregate_signals(workbooks, files),
    }


def _iter_files(root: Path, *, max_files: int) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    paths = sorted(
        (
            p for p in root.rglob("*")
            if p.is_file()
            and not any(part.startswith(".") for part in p.relative_to(root).parts)
        ),
        key=lambda p: str(p.relative_to(root)),
    )
    for path in paths:
        try:
            stat = path.stat()
        except OSError:
            continue
        out.append(
            {
                "relative_path": str(path.relative_to(root)),
                "suffix": path.suffix.lower(),
                "size_bytes": stat.st_size,
            }
        )
        if len(out) >= max_files:
            break
    return out


def _summarize_workbook(
    path: Path,
    root: Path,
    *,
    max_sheets: int,
    max_cells_per_sheet: int,
    max_formulas_per_sheet: int,
    max_scan_rows: int,
    max_scan_cols: int,
) -> dict[str, Any]:
    rel_path = str(path.relative_to(root))
    base: dict[str, Any] = {
        "relative_path": rel_path,
        "suffix": path.suffix.lower(),
        "size_bytes": path.stat().st_size if path.exists() else None,
        "readable": False,
        "sheets": [],
    }

    if path.suffix.lower() not in OPENPYXL_EXTENSIONS:
        base["error"] = "legacy .xls is detected but cannot be inspected with openpyxl"
        return base

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - exercised only without dependency
        base["error"] = f"openpyxl is required: {exc}"
        return base

    try:
        workbook = openpyxl.load_workbook(path, data_only=False)
    except Exception as exc:  # corrupted/unsupported workbook should become judge evidence
        base["error"] = f"failed to open workbook: {exc}"
        return base

    try:
        base["readable"] = True
        base["sheet_names"] = list(workbook.sheetnames)
        base["sheet_count"] = len(workbook.sheetnames)
        for sheet in list(workbook.worksheets)[:max_sheets]:
            base["sheets"].append(
                _summarize_sheet(
                    sheet,
                    max_cells=max_cells_per_sheet,
                    max_formulas=max_formulas_per_sheet,
                    max_scan_rows=max_scan_rows,
                    max_scan_cols=max_scan_cols,
                )
            )
        base["sheets_truncated"] = len(workbook.worksheets) > max_sheets
    finally:
        workbook.close()
    return base


def _summarize_sheet(
    sheet: Any,
    *,
    max_cells: int,
    max_formulas: int,
    max_scan_rows: int,
    max_scan_cols: int,
) -> dict[str, Any]:
    row_limit = min(sheet.max_row or 0, max_scan_rows)
    col_limit = min(sheet.max_column or 0, max_scan_cols)
    non_empty_samples: list[dict[str, Any]] = []
    formula_samples: list[dict[str, Any]] = []
    header_candidates: list[dict[str, Any]] = []
    number_formats: dict[str, int] = {}
    fill_colors: dict[str, int] = {}
    bold_count = 0
    numeric_count = 0
    text_count = 0
    blank_count = 0
    non_empty_count = 0

    for row in sheet.iter_rows(min_row=1, max_row=row_limit, min_col=1, max_col=col_limit):
        row_values: list[str] = []
        row_non_empty = 0
        for cell in row:
            value = cell.value
            if value is None:
                blank_count += 1
                continue
            non_empty_count += 1
            row_non_empty += 1
            display_value = _short_value(value)
            row_values.append(display_value)
            if len(non_empty_samples) < max_cells:
                non_empty_samples.append({"cell": cell.coordinate, "value": display_value})
            if isinstance(value, str) and value.startswith("="):
                if len(formula_samples) < max_formulas:
                    formula_samples.append({"cell": cell.coordinate, "formula": display_value})
            elif isinstance(value, (int, float)):
                numeric_count += 1
            else:
                text_count += 1
            if getattr(cell.font, "bold", False):
                bold_count += 1
            fmt = getattr(cell, "number_format", None)
            if fmt:
                number_formats[fmt] = number_formats.get(fmt, 0) + 1
            color = _cell_fill_rgb(cell)
            if color:
                fill_colors[color] = fill_colors.get(color, 0) + 1
        if row_non_empty >= 2 and len(header_candidates) < 3:
            header_candidates.append({"row": row[0].row, "values": row_values[:20]})

    return {
        "title": sheet.title,
        "max_row": sheet.max_row,
        "max_column": sheet.max_column,
        "scanned_rows": row_limit,
        "scanned_columns": col_limit,
        "non_empty_count_scanned": non_empty_count,
        "blank_count_scanned": blank_count,
        "numeric_count_scanned": numeric_count,
        "text_count_scanned": text_count,
        "formula_count_scanned": len(formula_samples),
        "non_empty_samples": non_empty_samples,
        "formula_samples": formula_samples,
        "formula_samples_truncated": len(formula_samples) >= max_formulas,
        "header_candidates": header_candidates,
        "merged_ranges": [str(rng) for rng in list(sheet.merged_cells.ranges)[:30]],
        "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
        "auto_filter_ref": sheet.auto_filter.ref,
        "table_names": list(getattr(sheet, "tables", {}).keys())[:30],
        "chart_count": len(getattr(sheet, "_charts", []) or []),
        "image_count": len(getattr(sheet, "_images", []) or []),
        "style_signals": {
            "bold_cell_count_scanned": bold_count,
            "top_number_formats": _top_counts(number_formats),
            "top_fill_colors": _top_counts(fill_colors),
        },
    }


def _aggregate_signals(workbooks: list[dict[str, Any]], files: list[dict[str, Any]]) -> dict[str, Any]:
    readable = [wb for wb in workbooks if wb.get("readable")]
    all_sheets = [sheet for wb in readable for sheet in wb.get("sheets", [])]
    formula_count = sum(int(sheet.get("formula_count_scanned", 0)) for sheet in all_sheets)
    chart_count = sum(int(sheet.get("chart_count", 0)) for sheet in all_sheets)
    image_count = sum(int(sheet.get("image_count", 0)) for sheet in all_sheets)
    table_count = sum(len(sheet.get("table_names", [])) for sheet in all_sheets)
    return {
        "has_readable_workbook": bool(readable),
        "unreadable_workbook_count": len(workbooks) - len(readable),
        "total_sheet_count": sum(int(wb.get("sheet_count", 0)) for wb in readable),
        "scanned_formula_count": formula_count,
        "has_formulas": formula_count > 0,
        "scanned_chart_count": chart_count,
        "scanned_image_count": image_count,
        "scanned_table_count": table_count,
        "has_non_excel_files": any(f["suffix"].lower() not in EXCEL_EXTENSIONS for f in files),
        "file_suffixes": sorted({f["suffix"].lower() or "<none>" for f in files}),
    }


def _short_value(value: Any, *, limit: int = 160) -> str:
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    if len(text) > limit:
        return text[: limit - 3] + "..."
    return text


def _cell_fill_rgb(cell: Any) -> str | None:
    fill = getattr(cell, "fill", None)
    fg = getattr(fill, "fgColor", None)
    rgb = getattr(fg, "rgb", None)
    if not rgb:
        return None
    rgb_s = str(rgb)
    if rgb_s in {"00000000", "000000"}:
        return None
    return rgb_s[-6:].upper()


def _top_counts(counts: dict[str, int], *, limit: int = 8) -> list[dict[str, Any]]:
    return [
        {"value": key, "count": value}
        for key, value in sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))[:limit]
    ]
